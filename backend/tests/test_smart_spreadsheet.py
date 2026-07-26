from datetime import date
from io import BytesIO

from openpyxl import Workbook

from app.smart_spreadsheet import parse_smart_workbook


def _multi_sheet_workbook() -> bytes:
    workbook = Workbook()
    gst = workbook.active
    gst.title = "GST Summary"
    gst.append(["GST register for April", None, None, None])
    gst.append(["Invoice Date", "State", "Taxable Value", "IGST"])
    gst.append([date(2026, 4, 1), "Gujarat", 1000, 180])
    gst.append([date(2026, 4, 2), "Maharashtra", 2000, 360])
    gst.append([date(2026, 5, 3), "Gujarat", 500, 90])
    gst.append([None, "Grand Total", 3500, 630])

    materials = workbook.create_sheet("Material Watch")
    materials.append(["Month", "Material", "Average Price", "Quantity"])
    materials.append(["Apr 2026", "Brass", 610, 40])
    materials.append(["May 2026", "Brass", 635, 35])
    materials.append(["Apr 2026", "Steel", 72, 120])
    materials.append(["May 2026", "Steel", 76, 140])

    hidden = workbook.create_sheet("Helper")
    hidden.sheet_state = "hidden"
    hidden.append(["Value", "Code"])
    hidden.append([1, "A"])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_smart_workbook_profiles_every_visible_tabular_sheet():
    parsed = parse_smart_workbook(_multi_sheet_workbook(), "operations.xlsx")

    assert len(parsed.datasets) == 2
    assert parsed.row_count == 7
    assert parsed.skipped_sheets == ("Helper",)

    gst = parsed.datasets[0]
    assert gst.sheet_name == "GST Summary"
    assert gst.domain == "gst"
    assert gst.header_row == 2
    assert gst.row_count == 3
    assert {column["label"] for column in gst.metric_columns} == {
        "Taxable Value",
        "IGST",
    }
    assert any(kpi["label"] == "Total Taxable Value" for kpi in gst.kpis)
    assert {chart["type"] for chart in gst.charts} >= {"line", "bar", "donut"}
    assert any("total row" in warning for warning in gst.warnings)

    materials = parsed.datasets[1]
    assert materials.domain == "inventory"
    assert materials.date_columns[0]["label"] == "Month"
    assert any(chart["id"] == "ranking" for chart in materials.charts)


def test_smart_csv_handles_unfamiliar_operational_schema():
    content = (
        "Recorded On,Machine,Output Units,Reject Units\n"
        "2026-04-01,Press A,1200,12\n"
        "2026-04-02,Press B,980,23\n"
        "2026-04-03,Press A,1310,10\n"
    ).encode()

    parsed = parse_smart_workbook(content, "factory-output.csv")

    dataset = parsed.datasets[0]
    assert dataset.domain == "general"
    assert dataset.row_count == 3
    assert {column["label"] for column in dataset.metric_columns} == {
        "Output Units",
        "Reject Units",
    }
    assert dataset.charts


def test_smart_workbook_keeps_and_flags_identical_rows():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Expenses"
    sheet.append(["Date", "Category", "Amount"])
    sheet.append([date(2026, 4, 1), "Freight", 500])
    sheet.append([date(2026, 4, 1), "Freight", 500])
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    parsed = parse_smart_workbook(output.getvalue(), "expenses.xlsx")

    assert parsed.datasets[0].duplicate_rows == 1
    assert parsed.datasets[0].row_count == 2
    assert any("possible duplicates" in warning for warning in parsed.datasets[0].warnings)
