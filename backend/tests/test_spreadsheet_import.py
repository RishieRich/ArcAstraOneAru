from datetime import date
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

from app.spreadsheet_import import (
    ImportValidationError,
    _source_key,
    parse_tally_workbook,
)
from app.routers.imports import _supplemental_smart_analysis


def _tally_workbook(kind: str) -> bytes:
    workbook = Workbook()
    vouchers = workbook.active
    vouchers.title = "Vouchers"
    vouchers.append(
        [
            "VoucherSeq",
            "VCHTYPE",
            "DATE",
            "GUID",
            "PARTYNAME",
            "VOUCHERNUMBER",
        ]
    )
    voucher_type = {"sales": "Sales", "purchase": "Purchase", "expense": "Journal"}[kind]
    vouchers.append([1, voucher_type, date(2026, 4, 5), f"{kind}-guid", "Test Party", "1"])

    ledgers = workbook.create_sheet("Ledger Entries")
    ledgers.append(
        [
            "VoucherSeq",
            "VoucherNumber",
            "VoucherType",
            "Date",
            "Party",
            "GUID",
            "LEDGERNAME",
            "ISDEEMEDPOSITIVE",
            "ISPARTYLEDGER",
            "AMOUNT",
        ]
    )
    if kind == "sales":
        rows = [
            [1, "1", voucher_type, date(2026, 4, 5), "Test Party", f"{kind}-guid",
             "Test Party", "Yes", "Yes", -118],
            [1, "1", voucher_type, date(2026, 4, 5), "Test Party", f"{kind}-guid",
             "Sales Ledger", "No", "No", 100],
            [1, "1", voucher_type, date(2026, 4, 5), "Test Party", f"{kind}-guid",
             "IGST", "No", "No", 18],
        ]
    else:
        category = "Purchase Ledger" if kind == "purchase" else "Repairs"
        rows = [
            [1, "1", voucher_type, date(2026, 4, 5), "Test Party", f"{kind}-guid",
             "Test Party", "No", "Yes", 118],
            [1, "1", voucher_type, date(2026, 4, 5), "Test Party", f"{kind}-guid",
             category, "Yes", "No", -100],
            [1, "1", voucher_type, date(2026, 4, 5), "Test Party", f"{kind}-guid",
             "IGST", "Yes", "No", -18],
        ]
    for row in rows:
        ledgers.append(row)

    if kind in {"sales", "purchase"}:
        inventory = workbook.create_sheet("Inventory Entries")
        inventory.append(
            [
                "VoucherSeq",
                "VoucherNumber",
                "VoucherType",
                "Date",
                "Party",
                "GUID",
                "STOCKITEMNAME",
                "RATE",
                "AMOUNT",
                "ACTUALQTY",
                "BILLEDQTY",
            ]
        )
        inventory.append(
            [
                1,
                "1",
                voucher_type,
                date(2026, 4, 5),
                "Test Party",
                f"{kind}-guid",
                "Widget",
                50,
                100 if kind == "sales" else -100,
                "2 Nos",
                "2 Nos",
            ]
        )

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _adaptive_sales_register(*, duplicate_business_row: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales Register"
    sheet.append(
        [
            "Date",
            "Particulars",
            "Buyer",
            "Voucher Type",
            "Quantity",
            "Alt. Units",
            "Rate",
            "Value",
            "Gross Total",
            "Interstate Sales",
            "IGST",
        ]
    )
    sheet.append(
        [
            date(2026, 4, 3),
            "Acme Customer",
            "Acme Customer",
            "Sales",
            10,
            None,
            None,
            100,
            118,
            100,
            18,
        ]
    )
    sheet.append([None, "Widget A", "", "", 10, None, 10, 100])
    if duplicate_business_row:
        sheet.append(
            [
                date(2026, 4, 3),
                "Acme Customer",
                "Acme Customer",
                "Sales",
                10,
                None,
                None,
                100,
                118,
                100,
                18,
            ]
        )
        sheet.append([None, "Widget A", "", "", 10, None, 10, 100])
    sheet.append([None, "Grand Total", "", "", None, None, None, 100, 118])

    # A second concatenated section shifts quantity into Alt. Units and puts a
    # custom party label where Voucher Type appeared in the first section.
    sheet.append(
        [
            date(2026, 5, 4),
            "Beta Customer",
            "Beta Customer",
            "Beta Customer",
            "Ahmedabad",
            5,
            None,
            50,
        ]
    )
    sheet.append([None, "Widget B", "", "", "", 5, 10, 50])
    sheet.append([None, "Grand Total", "", "", None, None, None, 75])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _profit_loss_statement() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Profit & Loss"
    sheet.append(["Profit & Loss A/c", None, None, None])
    sheet.append([None, "31-Mar-2027", None, "31-Mar-2027"])
    sheet.append(["Expenses", "Amount", "Income", "Amount"])
    sheet.append(["Purchase Accounts", 500, "Sales Accounts", 1000])
    sheet.append(["Rent", 100, "Other Income", 50])
    sheet.append(["Gross Profit", 450, "Grand Total", 1050])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@pytest.mark.parametrize("kind", ["sales", "purchase", "expense"])
def test_tally_workbook_is_classified_and_normalized(kind):
    result = parse_tally_workbook(_tally_workbook(kind), f"{kind}.xlsx", kind)

    assert result.detected_kind == kind
    assert result.min_date == date(2026, 4, 5)
    assert len(result.transactions) == 1
    transaction = result.transactions[0]
    assert transaction.source_key == f"guid:{kind}-guid"
    assert transaction.gross_amount == 118
    assert transaction.net_amount == 100
    assert transaction.tax_amount == 18
    assert transaction.category in {"Widget", "Repairs"}
    assert {line.line_type for line in transaction.lines} >= {"tax"}


def test_declared_type_mismatch_is_rejected_before_database_write():
    with pytest.raises(ImportValidationError, match="looks like sales"):
        parse_tally_workbook(_tally_workbook("sales"), "sales.xlsx", "purchase")


def test_legacy_xls_is_rejected_with_export_guidance():
    with pytest.raises(ImportValidationError, match=r"\.xlsx"):
        parse_tally_workbook(b"not a workbook", "sales.xls", "sales")


def test_fallback_voucher_identity_does_not_depend_on_export_sequence():
    first = _source_key(
        kind="sales",
        guid=None,
        sequence="1",
        txn_date=date(2026, 4, 5),
        voucher_number="INV-7",
        party="Test Party",
    )
    reordered = _source_key(
        kind="sales",
        guid=None,
        sequence="938",
        txn_date=date(2026, 4, 5),
        voucher_number=" inv-7 ",
        party="test party",
    )

    assert first == reordered


def test_repeated_voucher_guid_inside_one_workbook_is_counted_once():
    workbook = load_workbook(BytesIO(_tally_workbook("sales")))
    vouchers = workbook["Vouchers"]
    vouchers.append(
        [2, "Sales", date(2026, 4, 5), "sales-guid", "Test Party", "1"]
    )

    ledgers = workbook["Ledger Entries"]
    original_ledger_rows = list(ledgers.iter_rows(min_row=2, values_only=True))
    for row in original_ledger_rows:
        ledgers.append([2, *row[1:]])

    inventory = workbook["Inventory Entries"]
    original_inventory_rows = list(inventory.iter_rows(min_row=2, values_only=True))
    for row in original_inventory_rows:
        inventory.append([2, *row[1:]])

    output = BytesIO()
    workbook.save(output)
    workbook.close()

    parsed = parse_tally_workbook(output.getvalue(), "sales.xlsx", "sales")

    assert len(parsed.transactions) == 1
    assert parsed.skipped_rows == 1
    assert len(parsed.transactions[0].lines) == 2


def test_financial_import_endpoint_requires_dashboard_login(client):
    response = client.post(
        f"/v1/imports/financials?tenant_id={uuid4()}&declared_kind=sales",
        headers={"X-File-Name": "sales.xlsx"},
        content=b"not sent without auth",
    )
    assert response.status_code == 401


def test_adaptive_register_handles_concatenated_row_layouts_and_products():
    parsed = parse_tally_workbook(
        _adaptive_sales_register(),
        "Sale 26-27.xlsx",
        "sales",
    )

    assert parsed.source_format == "adaptive-flat-register"
    assert parsed.detected_kind == "sales"
    assert len(parsed.transactions) == 2
    assert parsed.column_mapping["quantity"] == "Quantity / Alt. Units"
    assert parsed.transactions[0].gross_amount == 118
    assert parsed.transactions[0].tax_amount == 18
    assert parsed.transactions[0].lines[0].name == "Widget A"
    assert parsed.transactions[0].lines[0].quantity == 10
    assert parsed.transactions[1].gross_amount == 50
    assert parsed.transactions[1].lines[0].name == "Widget B"
    assert parsed.transactions[1].lines[0].quantity == 5
    assert any("does not reconcile" in warning for warning in parsed.warnings)


def test_identical_flat_rows_without_voucher_number_are_flagged_not_dropped():
    parsed = parse_tally_workbook(
        _adaptive_sales_register(duplicate_business_row=True),
        "sales-register.xlsx",
        "sales",
    )

    matching = [
        transaction
        for transaction in parsed.transactions
        if transaction.party_name == "Acme Customer"
    ]
    assert len(matching) == 2
    assert len({transaction.source_key for transaction in matching}) == 2
    assert parsed.possible_duplicate_groups == 1
    assert any("kept and flagged" in warning for warning in parsed.warnings)


def test_adaptive_register_declared_type_mismatch_is_rejected():
    with pytest.raises(ImportValidationError, match="looks like sales"):
        parse_tally_workbook(
            _adaptive_sales_register(),
            "sales-register.xlsx",
            "purchase",
        )


def test_adaptive_register_accepts_tally_style_text_dates():
    workbook = load_workbook(BytesIO(_adaptive_sales_register()))
    workbook["Sales Register"]["A2"] = "3-Apr-26"
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    parsed = parse_tally_workbook(
        output.getvalue(),
        "sales-register.xlsx",
        "sales",
    )

    assert parsed.min_date == date(2026, 4, 3)


def test_profit_loss_summary_is_split_into_book_kinds_without_totals():
    parsed = parse_tally_workbook(
        _profit_loss_statement(),
        "Profit and Loss 26-27.xlsx",
        "auto",
    )

    assert parsed.detected_kind == "profit_loss"
    assert parsed.source_format == "adaptive-profit-loss"
    assert parsed.min_date == date(2027, 3, 31)
    assert {transaction.kind for transaction in parsed.transactions} == {
        "sales",
        "purchase",
        "expense",
    }
    assert sum(
        transaction.gross_amount
        for transaction in parsed.transactions
        if transaction.kind == "sales"
    ) == 1050
    assert sum(
        transaction.gross_amount
        for transaction in parsed.transactions
        if transaction.kind == "purchase"
    ) == 500
    assert sum(
        transaction.gross_amount
        for transaction in parsed.transactions
        if transaction.kind == "expense"
    ) == 100
    assert all("Gross Profit" != transaction.category for transaction in parsed.transactions)


def test_profit_loss_with_only_one_recognizable_side_is_rejected():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Profit & Loss"
    sheet.append(["Profit & Loss", "Amount"])
    sheet.append(["Sales Accounts", 1000])
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    with pytest.raises(ImportValidationError, match="only one side"):
        parse_tally_workbook(output.getvalue(), "pnl.xlsx", "profit_loss")


def test_known_finance_workbook_profiles_only_additional_useful_sheets():
    workbook = load_workbook(BytesIO(_tally_workbook("sales")))
    research = workbook.create_sheet("Material Watch")
    research.append(["Month", "Material", "Market Price"])
    research.append(["Apr 2026", "Brass", 610])
    research.append(["May 2026", "Brass", 635])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    data = output.getvalue()

    parsed = parse_tally_workbook(data, "sales-with-research.xlsx", "sales")
    supplemental = _supplemental_smart_analysis(
        data,
        "sales-with-research.xlsx",
        parsed,
    )

    assert supplemental is not None
    assert [dataset.sheet_name for dataset in supplemental.datasets] == ["Material Watch"]
