"""Deterministic, multi-sheet analytics for unfamiliar business workbooks.

This is the broad fallback behind "Smart Excel". It does not pretend that an
unknown table is a Sales or GST register. Instead it finds tabular regions,
infers typed columns, and produces an auditable profile of KPIs and charts.
Known Tally/finance layouts still pass through ``spreadsheet_import.py`` for
voucher-level accounting semantics.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import PurePath

import xlrd
from openpyxl import load_workbook

from app.spreadsheet_import import (
    ImportValidationError,
    MAX_UPLOAD_BYTES,
    clean_filename,
    _validate_xlsx_container,
)

MAX_SMART_SHEETS = 24
MAX_SMART_ROWS_PER_SHEET = 20_000
MAX_SMART_ROWS_PER_WORKBOOK = 40_000
MAX_SMART_COLUMNS = 120
MAX_CELL_TEXT = 500

_TOTAL_ROW = re.compile(r"^(grand\s+total|sub\s*total|subtotal|total)\b", re.I)
_ID_COLUMN = re.compile(
    r"\b(id|code|number|no|invoice|voucher|bill|phone|mobile|gstin|pan|hsn|sac|"
    r"account|reference|ref|serial|sr)\b",
    re.I,
)
_DATE_COLUMN = re.compile(
    r"\b(date|day|month|period|year|fy|financial\s+year|timestamp|time)\b",
    re.I,
)
_PERCENT_COLUMN = re.compile(
    r"\b(percent|percentage|pct|margin|rate|ratio|growth|share)\b|%",
    re.I,
)
_CURRENCY_COLUMN = re.compile(
    r"\b(amount|value|sales|revenue|purchase|expense|cost|profit|loss|tax|gst|"
    r"cgst|sgst|igst|utgst|cess|"
    r"balance|outstanding|receivable|payable|price|salary|wages|rent|freight|"
    r"turnover|income|debit|credit|total)\b",
    re.I,
)
_QUANTITY_COLUMN = re.compile(
    r"\b(qty|quantity|units?|volume|weight|stock|count)\b",
    re.I,
)

_DOMAIN_RULES = {
    "gst": (
        "gst",
        "gstin",
        "cgst",
        "sgst",
        "igst",
        "cess",
        "hsn",
        "sac",
        "taxable",
    ),
    "sales": (
        "sales",
        "sale",
        "revenue",
        "customer",
        "buyer",
        "invoice",
        "turnover",
    ),
    "purchase": (
        "purchase",
        "supplier",
        "vendor",
        "procurement",
        "material bought",
    ),
    "expense": (
        "expense",
        "spend",
        "cost",
        "payment",
        "rent",
        "salary",
        "wages",
        "freight",
        "travel",
        "overhead",
    ),
    "inventory": (
        "inventory",
        "stock",
        "sku",
        "item",
        "product",
        "material",
        "quantity",
        "warehouse",
    ),
    "payroll": (
        "payroll",
        "employee",
        "salary",
        "attendance",
        "department",
        "designation",
    ),
    "receivables": (
        "receivable",
        "outstanding",
        "overdue",
        "due date",
        "debtor",
        "collection",
    ),
}

_DOMAIN_TITLES = {
    "gst": "GST & tax data",
    "sales": "Sales data",
    "purchase": "Purchase data",
    "expense": "Expense data",
    "inventory": "Inventory & material data",
    "payroll": "Payroll data",
    "receivables": "Receivables data",
    "general": "Business data",
}


@dataclass(frozen=True)
class SmartRow:
    source_row: int
    row_fingerprint: str
    values: dict[str, object]


@dataclass(frozen=True)
class SmartDataset:
    sheet_name: str
    title: str
    domain: str
    confidence: float
    header_row: int
    row_count: int
    schema_fingerprint: str
    columns: tuple[dict, ...]
    date_columns: tuple[dict, ...]
    dimension_columns: tuple[dict, ...]
    metric_columns: tuple[dict, ...]
    kpis: tuple[dict, ...]
    charts: tuple[dict, ...]
    warnings: tuple[str, ...]
    duplicate_rows: int
    rows: tuple[SmartRow, ...]


@dataclass(frozen=True)
class SmartWorkbook:
    file_sha256: str
    filename: str
    datasets: tuple[SmartDataset, ...]
    skipped_sheets: tuple[str, ...]
    warnings: tuple[str, ...]

    @property
    def row_count(self) -> int:
        return sum(dataset.row_count for dataset in self.datasets)

    @property
    def duplicate_rows(self) -> int:
        return sum(dataset.duplicate_rows for dataset in self.datasets)


def parse_smart_workbook(data: bytes, filename: str) -> SmartWorkbook:
    """Profile every useful table in an xlsx/xlsm/csv upload."""
    safe_filename = clean_filename(filename)
    suffix = PurePath(safe_filename).suffix.casefold()
    if suffix not in {".xlsx", ".xlsm", ".xls", ".csv"}:
        raise ImportValidationError(
            "Smart Excel accepts .xlsx, .xlsm, legacy .xls and .csv files."
        )
    if not data:
        raise ImportValidationError("The selected file is empty.")
    if len(data) > MAX_UPLOAD_BYTES:
        raise ImportValidationError("The file is larger than the 5 MB upload limit.")

    file_sha256 = hashlib.sha256(data).hexdigest()
    tables: list[tuple[str, list[tuple[object, ...]], str]] = []
    skipped: list[str] = []
    workbook_warnings: list[str] = []

    if suffix == ".csv":
        tables.append((PurePath(safe_filename).stem or "CSV data", _csv_rows(data), "visible"))
    elif suffix == ".xls":
        try:
            legacy = xlrd.open_workbook(file_contents=data, on_demand=True)
        except (xlrd.XLRDError, OSError) as exc:
            raise ImportValidationError(
                "This legacy .xls workbook could not be read. Open it in Excel and "
                "save a fresh .xlsx copy."
            ) from exc
        try:
            for sheet in legacy.sheets()[:MAX_SMART_SHEETS]:
                rows = []
                for row_index in range(
                    min(sheet.nrows, MAX_SMART_ROWS_PER_SHEET + 30)
                ):
                    values = []
                    for column_index in range(min(sheet.ncols, MAX_SMART_COLUMNS)):
                        cell = sheet.cell(row_index, column_index)
                        value = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            try:
                                value = xlrd.xldate_as_datetime(
                                    value,
                                    legacy.datemode,
                                )
                            except (ValueError, OverflowError):
                                pass
                        values.append(value)
                    rows.append(tuple(values))
                tables.append((sheet.name, rows, "visible"))
            if legacy.nsheets > MAX_SMART_SHEETS:
                workbook_warnings.append(
                    f"Only the first {MAX_SMART_SHEETS} sheets were inspected."
                )
        finally:
            legacy.release_resources()
    else:
        _validate_xlsx_container(data)
        try:
            workbook = load_workbook(
                BytesIO(data),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as exc:
            raise ImportValidationError(
                "This workbook could not be read. Open it in Excel and save a fresh .xlsx copy."
            ) from exc
        try:
            for sheet in workbook.worksheets[:MAX_SMART_SHEETS]:
                state = getattr(sheet, "sheet_state", "visible")
                if state != "visible":
                    skipped.append(sheet.title)
                    continue
                rows = [
                    tuple(row)
                    for row in sheet.iter_rows(
                        min_row=1,
                        max_row=min(sheet.max_row or 1, MAX_SMART_ROWS_PER_SHEET + 30),
                        max_col=min(sheet.max_column or 1, MAX_SMART_COLUMNS),
                        values_only=True,
                    )
                ]
                tables.append((sheet.title, rows, state))
            if len(workbook.worksheets) > MAX_SMART_SHEETS:
                workbook_warnings.append(
                    f"Only the first {MAX_SMART_SHEETS} sheets were inspected."
                )
        finally:
            workbook.close()

    datasets: list[SmartDataset] = []
    total_rows = 0
    for sheet_name, raw_rows, _state in tables:
        if total_rows >= MAX_SMART_ROWS_PER_WORKBOOK:
            skipped.append(sheet_name)
            continue
        remaining = MAX_SMART_ROWS_PER_WORKBOOK - total_rows
        dataset = _profile_table(sheet_name, raw_rows, remaining)
        if dataset is None:
            skipped.append(sheet_name)
            continue
        datasets.append(dataset)
        total_rows += dataset.row_count

    if not datasets:
        raise ImportValidationError(
            "ARQ could not find a usable table. Keep one header row above the data and "
            "include at least two populated columns."
        )
    if skipped:
        workbook_warnings.append(
            f"{len(skipped)} empty, hidden or non-tabular sheet(s) were skipped."
        )
    if total_rows >= MAX_SMART_ROWS_PER_WORKBOOK:
        workbook_warnings.append(
            f"Analysis stopped at the safe {MAX_SMART_ROWS_PER_WORKBOOK:,}-row workbook limit."
        )
    return SmartWorkbook(
        file_sha256=file_sha256,
        filename=safe_filename,
        datasets=tuple(datasets),
        skipped_sheets=tuple(skipped),
        warnings=tuple(workbook_warnings),
    )


def _csv_rows(data: bytes) -> list[tuple[object, ...]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ImportValidationError("The CSV text encoding could not be read.")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [
        tuple(row[:MAX_SMART_COLUMNS])
        for row in csv.reader(StringIO(text), dialect)
    ][: MAX_SMART_ROWS_PER_SHEET + 30]


def _profile_table(
    sheet_name: str,
    raw_rows: list[tuple[object, ...]],
    remaining_rows: int,
) -> SmartDataset | None:
    header_index = _find_header_row(raw_rows)
    if header_index is None:
        return None
    raw_header = raw_rows[header_index]
    last_column = max(
        (index for index, value in enumerate(raw_header) if _present(value)),
        default=-1,
    )
    if last_column < 1:
        return None
    width = min(last_column + 1, MAX_SMART_COLUMNS)
    headers = _unique_headers(raw_header[:width])
    source_rows: list[tuple[int, tuple[object, ...]]] = []
    skipped_totals = 0
    blank_run = 0
    for source_row, values in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        row = tuple(values[:width]) + (None,) * max(0, width - len(values))
        if not any(_present(value) for value in row):
            blank_run += 1
            if blank_run >= 40 and source_rows:
                break
            continue
        blank_run = 0
        first_text = next((_text(value) for value in row if _text(value)), "")
        if first_text and _TOTAL_ROW.search(first_text):
            skipped_totals += 1
            continue
        source_rows.append((source_row, row))
        if len(source_rows) >= min(MAX_SMART_ROWS_PER_SHEET, remaining_rows):
            break
    if len(source_rows) < 2:
        return None

    column_profiles = _infer_columns(headers, [row for _, row in source_rows])
    metrics = [column for column in column_profiles if column["role"] == "metric"]
    dates = [column for column in column_profiles if column["role"] == "date"]
    dimensions = [column for column in column_profiles if column["role"] == "dimension"]
    if not metrics and not dimensions:
        return None

    normalized_rows = _normalize_rows(headers, column_profiles, source_rows)
    fingerprints = Counter(row.row_fingerprint for row in normalized_rows)
    duplicate_rows = sum(count - 1 for count in fingerprints.values() if count > 1)
    domain, domain_confidence = _classify_domain(sheet_name, headers)
    title = (
        f"{_DOMAIN_TITLES[domain]} · {sheet_name}"
        if sheet_name.casefold() not in _DOMAIN_TITLES[domain].casefold()
        else _DOMAIN_TITLES[domain]
    )
    kpis = _build_kpis(column_profiles, normalized_rows)
    charts = _build_charts(
        sheet_name=sheet_name,
        columns=column_profiles,
        rows=normalized_rows,
    )
    warnings = []
    if skipped_totals:
        warnings.append(
            f"{skipped_totals} printed total row(s) were excluded to avoid double-counting."
        )
    if duplicate_rows:
        warnings.append(
            f"{duplicate_rows} repeated row(s) were kept and marked as possible duplicates."
        )
    if not metrics:
        warnings.append(
            "No reliable numeric measure was found; ARQ is showing record and category counts."
        )
    if not dates:
        warnings.append(
            "No reliable date column was found, so time-trend charts are unavailable."
        )
    if len(source_rows) >= min(MAX_SMART_ROWS_PER_SHEET, remaining_rows):
        warnings.append("This sheet reached the safe row-analysis limit.")

    typed_share = sum(
        column["coverage"] for column in column_profiles if column["role"] in {"metric", "date"}
    ) / max(len(column_profiles), 1)
    confidence = min(
        0.97,
        max(0.55, 0.52 + domain_confidence * 0.2 + min(typed_share, 1) * 0.2),
    )
    schema_seed = json.dumps(
        {
            "sheet": _key(sheet_name),
            "columns": [(column["key"], column["type"], column["role"]) for column in column_profiles],
        },
        sort_keys=True,
    ).encode("utf-8")
    return SmartDataset(
        sheet_name=sheet_name[:120],
        title=title[:180],
        domain=domain,
        confidence=round(confidence, 3),
        header_row=header_index + 1,
        row_count=len(normalized_rows),
        schema_fingerprint=hashlib.sha256(schema_seed).hexdigest(),
        columns=tuple(column_profiles),
        date_columns=tuple(_public_column(column) for column in dates),
        dimension_columns=tuple(_public_column(column) for column in dimensions[:6]),
        metric_columns=tuple(_public_column(column) for column in metrics[:12]),
        kpis=tuple(kpis),
        charts=tuple(charts),
        warnings=tuple(warnings),
        duplicate_rows=duplicate_rows,
        rows=tuple(normalized_rows),
    )


def _find_header_row(rows: list[tuple[object, ...]]) -> int | None:
    best: tuple[float, int] | None = None
    for index, values in enumerate(rows[:30]):
        populated = [(column, value) for column, value in enumerate(values) if _present(value)]
        if len(populated) < 2 or len(populated) > MAX_SMART_COLUMNS:
            continue
        text_values = [_text(value) for _, value in populated]
        text_count = sum(bool(value) for value in text_values)
        if text_count < max(2, len(populated) // 2):
            continue
        normalized = [_key(value or "") for value in text_values if value]
        unique_share = len(set(normalized)) / max(len(normalized), 1)
        evidence_rows = rows[index + 1 : index + 9]
        evidence = sum(
            1
            for row in evidence_rows
            if sum(_present(row[column]) for column, _ in populated if column < len(row)) >= 2
        )
        if evidence < 2:
            continue
        score = len(populated) * 2 + text_count + unique_share * 4 + evidence * 1.5
        candidate = (score, -index)
        if best is None or candidate > (best[0], -best[1]):
            best = (score, index)
    return best[1] if best else None


def _unique_headers(values: tuple[object, ...]) -> list[dict[str, str]]:
    seen: Counter[str] = Counter()
    headers = []
    for index, value in enumerate(values):
        label = _text(value) or f"Column {index + 1}"
        base = _key(label) or f"column_{index + 1}"
        seen[base] += 1
        key = base if seen[base] == 1 else f"{base}_{seen[base]}"
        headers.append({"key": key, "label": label[:120]})
    return headers


def _infer_columns(
    headers: list[dict[str, str]],
    rows: list[tuple[object, ...]],
) -> list[dict]:
    profiles = []
    for index, header in enumerate(headers):
        values = [row[index] for row in rows if index < len(row) and _present(row[index])]
        total = len(values)
        date_values = [_date_value(value) for value in values]
        number_values = [_number_value(value) for value in values]
        date_count = sum(value is not None for value in date_values)
        number_count = sum(value is not None for value in number_values)
        label = header["label"]
        date_hint = bool(_DATE_COLUMN.search(label))
        id_hint = bool(_ID_COLUMN.search(label))
        if total and (
            date_count / total >= 0.7
            or (date_hint and date_count / total >= 0.45)
        ):
            value_type = "date"
            role = "date"
            coverage = date_count / total
        elif total and number_count / total >= 0.7 and not id_hint:
            value_type = "number"
            role = "metric"
            coverage = number_count / total
        else:
            value_type = "text"
            distinct = len({_display_value(value) for value in values})
            distinct_share = distinct / max(total, 1)
            role = (
                "identifier"
                if id_hint and distinct_share > 0.55
                else "dimension"
                if total and (distinct <= 80 or distinct_share <= 0.7)
                else "text"
            )
            coverage = total / max(len(rows), 1)
        value_format = _metric_format(label) if role == "metric" else value_type
        profiles.append(
            {
                "key": header["key"],
                "label": label,
                "type": value_type,
                "role": role,
                "format": value_format,
                "coverage": round(coverage, 3),
                "distinct": len({_display_value(value) for value in values}),
            }
        )
    return profiles


def _normalize_rows(
    headers: list[dict[str, str]],
    columns: list[dict],
    source_rows: list[tuple[int, tuple[object, ...]]],
) -> list[SmartRow]:
    normalized = []
    for source_row, values in source_rows:
        output: dict[str, object] = {}
        for index, header in enumerate(headers):
            value = values[index] if index < len(values) else None
            column = columns[index]
            if not _present(value):
                output[header["key"]] = None
            elif column["type"] == "number":
                number = _number_value(value)
                output[header["key"]] = float(number) if number is not None else None
            elif column["type"] == "date":
                parsed_date = _date_value(value)
                output[header["key"]] = parsed_date.isoformat() if parsed_date else _display_value(value)
            else:
                output[header["key"]] = _display_value(value)
        fingerprint = hashlib.sha256(
            json.dumps(output, sort_keys=True, ensure_ascii=False).encode("utf-8")
        ).hexdigest()
        normalized.append(
            SmartRow(
                source_row=source_row,
                row_fingerprint=fingerprint,
                values=output,
            )
        )
    return normalized


def _build_kpis(columns: list[dict], rows: list[SmartRow]) -> list[dict]:
    metrics = [column for column in columns if column["role"] == "metric"]
    dimensions = [column for column in columns if column["role"] == "dimension"]
    ranked_metrics = sorted(metrics, key=_metric_priority, reverse=True)
    kpis = [
        {
            "key": "records",
            "label": "Records understood",
            "label_key": "records",
            "source_label": None,
            "value": len(rows),
            "format": "integer",
            "aggregation": "count",
        }
    ]
    for column in ranked_metrics[:3]:
        values = [
            row.values.get(column["key"])
            for row in rows
            if isinstance(row.values.get(column["key"]), (int, float))
        ]
        if not values:
            continue
        aggregate = "average" if column["format"] == "percent" or "rate" in column["key"] else "sum"
        value = sum(values) / len(values) if aggregate == "average" else sum(values)
        kpis.append(
            {
                "key": column["key"],
                "label": (
                    f"Average {column['label']}"
                    if aggregate == "average"
                    else f"Total {column['label']}"
                ),
                "label_key": aggregate,
                "source_label": column["label"],
                "value": round(value, 4),
                "format": column["format"],
                "aggregation": aggregate,
            }
        )
    if len(kpis) < 4 and dimensions:
        dimension = dimensions[0]
        distinct = {
            row.values.get(dimension["key"])
            for row in rows
            if row.values.get(dimension["key"]) not in (None, "")
        }
        kpis.append(
            {
                "key": f"distinct_{dimension['key']}",
                "label": f"Unique {dimension['label']}",
                "label_key": "distinct",
                "source_label": dimension["label"],
                "value": len(distinct),
                "format": "integer",
                "aggregation": "distinct",
            }
        )
    return kpis[:4]


def _build_charts(
    *,
    sheet_name: str,
    columns: list[dict],
    rows: list[SmartRow],
) -> list[dict]:
    metrics = sorted(
        [column for column in columns if column["role"] == "metric"],
        key=_metric_priority,
        reverse=True,
    )
    dates = [column for column in columns if column["role"] == "date"]
    dimensions = sorted(
        [column for column in columns if column["role"] == "dimension"],
        key=lambda column: (
            bool(re.search(r"\b(category|party|customer|vendor|product|item|material|state|type)\b", column["label"], re.I)),
            -column["distinct"],
        ),
        reverse=True,
    )
    charts = []
    primary_metric = metrics[0] if metrics else None

    if dates and primary_metric:
        grouped: defaultdict[str, float] = defaultdict(float)
        for row in rows:
            raw_date = row.values.get(dates[0]["key"])
            value = row.values.get(primary_metric["key"])
            if not raw_date or not isinstance(value, (int, float)):
                continue
            label = str(raw_date)[:7]
            grouped[label] += value
        points = [
            {"label": label, "value": round(value, 4)}
            for label, value in sorted(grouped.items())[-18:]
        ]
        if len(points) >= 2:
            charts.append(
                {
                    "id": "trend",
                    "type": "line",
                    "title": f"{primary_metric['label']} over time",
                    "subtitle": f"Grouped monthly from {dates[0]['label']}",
                    "metric_label": primary_metric["label"],
                    "dimension_label": dates[0]["label"],
                    "format": primary_metric["format"],
                    "points": points,
                }
            )

    if dimensions:
        dimension = dimensions[0]
        grouped: defaultdict[str, float] = defaultdict(float)
        for row in rows:
            label = row.values.get(dimension["key"])
            if label in (None, ""):
                continue
            value = row.values.get(primary_metric["key"]) if primary_metric else 1
            if primary_metric and not isinstance(value, (int, float)):
                continue
            grouped[str(label)[:80]] += float(value)
        points = [
            {"label": label, "value": round(value, 4)}
            for label, value in sorted(grouped.items(), key=lambda item: item[1], reverse=True)[:8]
        ]
        if points:
            charts.append(
                {
                    "id": "ranking",
                    "type": "bar",
                    "title": (
                        f"Top {dimension['label']} by {primary_metric['label']}"
                        if primary_metric
                        else f"Records by {dimension['label']}"
                    ),
                    "subtitle": f"Strongest contributors in {sheet_name}",
                    "metric_label": primary_metric["label"] if primary_metric else None,
                    "dimension_label": dimension["label"],
                    "format": primary_metric["format"] if primary_metric else "integer",
                    "points": points,
                }
            )

    comparable = [
        metric
        for metric in metrics
        if metric["format"] == "currency" or _CURRENCY_COLUMN.search(metric["label"])
    ][:6]
    if len(comparable) >= 2:
        points = []
        for metric in comparable:
            value = sum(
                row.values.get(metric["key"], 0)
                for row in rows
                if isinstance(row.values.get(metric["key"]), (int, float))
            )
            if value:
                points.append({"label": metric["label"], "value": round(value, 4)})
        if len(points) >= 2:
            charts.append(
                {
                    "id": "mix",
                    "type": "donut",
                    "title": "Metric mix",
                    "subtitle": "Comparable value columns from the uploaded sheet",
                    "metric_label": None,
                    "dimension_label": None,
                    "format": "currency",
                    "points": points,
                }
            )
    return charts[:3]


def _classify_domain(sheet_name: str, headers: list[dict[str, str]]) -> tuple[str, float]:
    haystack = " ".join([sheet_name, *(header["label"] for header in headers)]).casefold()
    scores = {
        domain: sum(1 for keyword in keywords if keyword in haystack)
        for domain, keywords in _DOMAIN_RULES.items()
    }
    best_domain, score = max(scores.items(), key=lambda item: item[1])
    if score == 0:
        return "general", 0.25
    second = sorted(scores.values(), reverse=True)[1]
    confidence = min(1.0, 0.45 + score * 0.12 + max(score - second, 0) * 0.08)
    return best_domain, confidence


def _metric_priority(column: dict) -> tuple[int, float, int]:
    label = column["label"]
    semantic = (
        5
        if _CURRENCY_COLUMN.search(label)
        else 4
        if _QUANTITY_COLUMN.search(label)
        else 3
        if _PERCENT_COLUMN.search(label)
        else 1
    )
    return semantic, column["coverage"], column["distinct"]


def _metric_format(label: str) -> str:
    if _PERCENT_COLUMN.search(label):
        return "percent"
    if _CURRENCY_COLUMN.search(label):
        return "currency"
    if _QUANTITY_COLUMN.search(label):
        return "number"
    return "number"


def _public_column(column: dict) -> dict:
    return {
        "key": column["key"],
        "label": column["label"],
        "type": column["type"],
        "role": column["role"],
        "format": column["format"],
        "coverage": column["coverage"],
        "distinct": column["distinct"],
    }


def _number_value(value: object) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    cleaned = (
        text.replace(",", "")
        .replace("₹", "")
        .replace("Rs.", "")
        .replace("Rs", "")
        .replace("%", "")
        .strip("() ")
    )
    try:
        number = Decimal(cleaned)
    except InvalidOperation:
        return None
    return -number if negative else number


def _date_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float, Decimal, bool)) or value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for pattern in (
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d.%m.%Y",
        "%d-%b-%Y",
        "%d-%b-%y",
        "%b %Y",
        "%B %Y",
    ):
        try:
            parsed = datetime.strptime(text, pattern)
            return parsed.date()
        except ValueError:
            continue
    return None


def _display_value(value: object) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()[:MAX_CELL_TEXT]


def _text(value: object) -> str:
    if isinstance(value, str):
        return value.strip()[:MAX_CELL_TEXT]
    return ""


def _present(value: object) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _key(value: str) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", value.casefold()).strip("_")
    return key[:80]
