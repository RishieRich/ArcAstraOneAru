"""Turn Tally-style Excel exports into a small, auditable finance model.

The importer deliberately uses workbook structure and voucher semantics rather
than asking an LLM to guess from customer data. That keeps uploads deterministic,
keeps party names and amounts on our server, and makes a wrong classification a
validation error instead of a silent database write.
"""
from __future__ import annotations

import hashlib
import json
import re
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import PurePath
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_ZIP_MEMBERS = 600
MAX_SHEET_ROWS = 100_000
MAX_SHEET_COLUMNS = 400
ALLOWED_KINDS = {"sales", "purchase", "expense"}
ALLOWED_IMPORT_KINDS = ALLOWED_KINDS | {"profit_loss"}

_NUMBER = re.compile(r"-?\d+(?:\.\d+)?")
_TAX_LEDGER = re.compile(r"(^|[^A-Z])(I?GST|CGST|SGST|UTGST|CESS|TAX)([^A-Z]|$)", re.I)
_ROUNDING_LEDGER = re.compile(r"round(?:ing)?[\s._-]*off", re.I)
_PROFIT_LOSS_TITLE = re.compile(
    r"\b(profit\s*(?:&|and)\s*loss|p\s*&\s*l|income\s+statement)\b",
    re.I,
)
_PROFIT_LOSS_SKIP = re.compile(
    r"^(particulars?|amount|closing\s+stock|opening\s+stock|"
    r"(grand\s+)?total|gross\s+(profit|loss)|net\s+(profit|loss)|"
    r"profit\s*(?:&|and)\s*loss(\s+a/?c)?)\b",
    re.I,
)
_PROFIT_LOSS_SALES = re.compile(
    r"\b(sales?|revenue|turnover|service\s+income|operating\s+income|"
    r"direct\s+income|indirect\s+income|other\s+income)\b",
    re.I,
)
_PROFIT_LOSS_PURCHASE = re.compile(
    r"\b(purchases?|cost\s+of\s+goods|cogs|material\s+consum|raw\s+material|"
    r"cost\s+of\s+sales)\b",
    re.I,
)
_PROFIT_LOSS_EXPENSE = re.compile(
    r"\b(expenses?|overheads?|salary|salaries|wages?|rent|freight|"
    r"depreciation|interest|commission|advertis|repairs?|electricity|"
    r"telephone|travell?ing|insurance|bank\s+charges?|professional\s+fees?)\b",
    re.I,
)


class ImportValidationError(ValueError):
    """A safe, user-fixable workbook error."""


@dataclass(frozen=True)
class ParsedLine:
    line_type: str
    name: str
    amount: Decimal
    quantity: Decimal | None = None
    unit: str | None = None
    rate: Decimal | None = None


@dataclass(frozen=True)
class ParsedTransaction:
    source_key: str
    source_row: int
    kind: str
    txn_date: date | None
    voucher_number: str | None
    voucher_type: str | None
    party_name: str | None
    category: str | None
    gross_amount: Decimal
    net_amount: Decimal
    tax_amount: Decimal
    lines: tuple[ParsedLine, ...]


@dataclass(frozen=True)
class ParsedWorkbook:
    file_sha256: str
    detected_kind: str
    confidence: float
    classification_reason: str
    transactions: tuple[ParsedTransaction, ...]
    skipped_rows: int
    min_date: date | None
    max_date: date | None
    source_format: str
    column_mapping: dict[str, str]
    warnings: tuple[str, ...]
    possible_duplicate_groups: int

    @property
    def line_count(self) -> int:
        return sum(len(transaction.lines) for transaction in self.transactions)


def clean_filename(value: str) -> str:
    """Keep a display-only basename; never trust a browser-provided path."""
    name = PurePath(value.replace("\\", "/")).name
    name = "".join(ch for ch in name if ch.isprintable() and ch not in "\r\n\t").strip()
    return (name or "upload.xlsx")[:180]


def parse_tally_workbook(
    data: bytes,
    filename: str,
    declared_kind: str | None = None,
) -> ParsedWorkbook:
    declared = _normalize_kind(declared_kind)
    safe_filename = clean_filename(filename)
    if not safe_filename.lower().endswith(".xlsx"):
        raise ImportValidationError(
            "Please export or save the file as .xlsx. Legacy .xls files are not supported."
        )
    if not data:
        raise ImportValidationError("The selected workbook is empty.")
    if len(data) > MAX_UPLOAD_BYTES:
        raise ImportValidationError("The workbook is larger than the 5 MB upload limit.")

    _validate_xlsx_container(data)
    file_sha256 = hashlib.sha256(data).hexdigest()

    try:
        workbook = load_workbook(
            BytesIO(data),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ImportValidationError(
            "This file could not be read as an Excel workbook. Re-export it from Tally as .xlsx."
        ) from exc

    try:
        vouchers_sheet = _find_sheet(workbook, "vouchers")
        ledgers_sheet = _find_sheet(workbook, "ledgerentries")
        inventory_sheet = _find_sheet(workbook, "inventoryentries")
        if ledgers_sheet is not None:
            vouchers = _records(vouchers_sheet) if vouchers_sheet is not None else []
            ledgers = _records(ledgers_sheet)
            inventory = _records(inventory_sheet) if inventory_sheet is not None else []

            detected_kind, confidence, reason = _classify(
                safe_filename,
                vouchers,
                ledgers,
                declared,
            )
            transactions, skipped_rows = _build_transactions(
                detected_kind,
                vouchers,
                ledgers,
                inventory,
            )
            source_format = "tally-multi-sheet"
            column_mapping = {
                "transactions": vouchers_sheet.title if vouchers_sheet else "Ledger Entries",
                "ledger_lines": ledgers_sheet.title,
                "product_lines": inventory_sheet.title if inventory_sheet else "Not supplied",
            }
            warnings: tuple[str, ...] = ()
            possible_duplicate_groups = 0
        else:
            profit_loss_sheet = _find_profit_loss_sheet(
                workbook,
                force=declared == "profit_loss" or _filename_is_profit_loss(safe_filename),
            )
            if profit_loss_sheet is not None:
                transactions, skipped_rows, period_columns = (
                    _build_profit_loss_transactions(profit_loss_sheet)
                )
                detected_kind = "profit_loss"
                confidence = 0.9
                reason = (
                    f"Profit & Loss summary detected on '{profit_loss_sheet.title}' and "
                    "normalized into sales, purchase and expense ledger categories."
                )
                source_format = "adaptive-profit-loss"
                column_mapping = {
                    "sheet": profit_loss_sheet.title,
                    "labels": "Particular / ledger text",
                    "amounts": "Nearest numeric value to each ledger",
                    "period_columns": str(period_columns),
                }
                warnings = (
                    "Profit & Loss summary rows were imported as ledger categories; "
                    "voucher, customer and product-level detail is not available from this file.",
                    "Printed totals and Gross/Net Profit or Loss rows were skipped to avoid "
                    "double-counting the underlying ledgers.",
                    "Opening and closing stock rows were skipped because their accounting "
                    "treatment cannot be inferred safely from an arbitrary workbook layout.",
                )
                possible_duplicate_groups = 0
            else:
                flat = _find_flat_register(workbook)
                if flat is None:
                    raise ImportValidationError(
                        "No supported transaction table was found. Include either Tally's "
                        "'Ledger Entries' sheet, a register with Date, Particulars and Value/"
                        "Amount columns, or a clearly titled Profit & Loss statement."
                    )
                sheet, header_row, mapping = flat
                detected_kind, confidence, reason = _classify_flat_register(
                    sheet,
                    header_row,
                    mapping,
                    safe_filename,
                    declared,
                )
                (
                    transactions,
                    skipped_rows,
                    possible_duplicate_groups,
                    reconciliation_warnings,
                ) = _build_flat_register_transactions(
                    sheet,
                    header_row,
                    mapping,
                    detected_kind,
                )
                source_format = "adaptive-flat-register"
                column_mapping = mapping["display"]
                warning_list = [
                    "A single-sheet register was detected; parent vouchers and product detail "
                    "rows were reconciled by date, value, quantity and rate."
                ]
                warning_list.extend(reconciliation_warnings)
                if mapping.get("voucher_number") is None:
                    warning_list.append(
                        "No voucher number or GUID column was found. Stable business fingerprints "
                        "are used to match re-uploads."
                    )
                if possible_duplicate_groups:
                    warning_list.append(
                        f"{possible_duplicate_groups} identical-looking transaction group(s) were "
                        "kept and flagged because the workbook has no reliable voucher identity."
                    )
                warnings = tuple(warning_list)
    finally:
        workbook.close()

    if not transactions:
        raise ImportValidationError(
            "The workbook has no usable voucher amounts. Include party and amount fields in "
            "the Tally export, then try again."
        )

    dated = [transaction.txn_date for transaction in transactions if transaction.txn_date]
    return ParsedWorkbook(
        file_sha256=file_sha256,
        detected_kind=detected_kind,
        confidence=confidence,
        classification_reason=reason,
        transactions=tuple(transactions),
        skipped_rows=skipped_rows,
        min_date=min(dated) if dated else None,
        max_date=max(dated) if dated else None,
        source_format=source_format,
        column_mapping=column_mapping,
        warnings=warnings,
        possible_duplicate_groups=possible_duplicate_groups,
    )


def _validate_xlsx_container(data: bytes) -> None:
    try:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ZIP_MEMBERS:
                raise ImportValidationError("The workbook contains too many internal files.")
            if sum(member.file_size for member in members) > MAX_UNCOMPRESSED_BYTES:
                raise ImportValidationError(
                    "The expanded workbook is too large to process safely."
                )
            names = {member.filename for member in members}
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise ImportValidationError("The selected file is not a valid .xlsx workbook.")
    except ImportValidationError:
        raise
    except (zipfile.BadZipFile, OSError) as exc:
        raise ImportValidationError("The selected file is not a valid .xlsx workbook.") from exc


def _find_sheet(workbook, normalized_name: str):
    for sheet in workbook.worksheets:
        if _header_key(sheet.title) == normalized_name.upper():
            return sheet
    return None


def _records(sheet) -> list[dict[str, object]]:
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        return []
    if len(raw_headers) > MAX_SHEET_COLUMNS:
        raise ImportValidationError(
            f"The '{sheet.title}' sheet has too many columns to process safely."
        )

    headers = [_header_key(value) for value in raw_headers]
    if not any(headers):
        raise ImportValidationError(f"The '{sheet.title}' sheet has no header row.")

    records: list[dict[str, object]] = []
    for row_number, values in enumerate(rows, start=2):
        if row_number > MAX_SHEET_ROWS + 1:
            raise ImportValidationError(
                f"The '{sheet.title}' sheet exceeds the {MAX_SHEET_ROWS:,}-row limit."
            )
        record = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in record.values()):
            record["_ROW"] = row_number
            records.append(record)
    return records


def _find_profit_loss_sheet(workbook, *, force: bool = False):
    """Find a visibly labelled P&L without treating an arbitrary total as one."""
    fallback = None
    for sheet in workbook.worksheets:
        visible_text: list[str] = []
        has_material_value = False
        for values in sheet.iter_rows(
            min_row=1,
            max_row=min(sheet.max_row or 1, 18),
            max_col=min(sheet.max_column or 1, 16),
            values_only=True,
        ):
            for value in values:
                text = _text(value)
                if text:
                    visible_text.append(text)
                if _looks_like_amount(value) and abs(_amount(value)) > 0:
                    has_material_value = True
        joined = " ".join([sheet.title, *visible_text])
        if _PROFIT_LOSS_TITLE.search(joined) and has_material_value:
            return sheet
        if fallback is None and has_material_value and visible_text:
            fallback = sheet
    return fallback if force else None


def _build_profit_loss_transactions(sheet) -> tuple[list[ParsedTransaction], int, int]:
    """Normalize common one- and two-sided P&L layouts into category rows.

    Arbitrary statements are deliberately handled conservatively: only a label
    with a nearby numeric value and a recognizable/section-derived book kind is
    accepted. Printed totals are excluded because importing both them and their
    child ledgers would double the result.
    """
    period_by_column = _profit_loss_periods(sheet)
    section_by_column: dict[int, str] = {}
    occurrences: Counter[str] = Counter()
    parsed: list[ParsedTransaction] = []
    skipped = 0

    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number > MAX_SHEET_ROWS:
            raise ImportValidationError(
                f"The '{sheet.title}' sheet exceeds the {MAX_SHEET_ROWS:,}-row limit."
            )
        if len(values) > MAX_SHEET_COLUMNS:
            raise ImportValidationError(
                f"The '{sheet.title}' sheet has too many columns to process safely."
            )

        label_columns = [
            index
            for index, value in enumerate(values)
            if _text(value) and not _looks_like_amount(value)
        ]
        for label_position, label_column in enumerate(label_columns):
            label = _text(values[label_column])
            if not label:
                continue
            next_label = (
                label_columns[label_position + 1]
                if label_position + 1 < len(label_columns)
                else len(values)
            )
            amount_column = next(
                (
                    index
                    for index in range(label_column + 1, next_label)
                    if _looks_like_amount(values[index])
                    and abs(_amount(values[index])) > 0
                ),
                None,
            )

            section = _profit_loss_kind(label)
            if amount_column is None:
                if section:
                    section_by_column[label_column] = section
                continue
            if _PROFIT_LOSS_SKIP.search(label):
                skipped += 1
                continue

            inherited = _nearest_section(section_by_column, label_column)
            kind = section or inherited
            if kind not in ALLOWED_KINDS:
                skipped += 1
                continue

            amount = _money(abs(_amount(values[amount_column])))
            txn_date = period_by_column.get(amount_column)
            stable_label = _header_key(label)
            occurrence_key = f"{kind}\x1f{txn_date or ''}\x1f{stable_label}"
            occurrences[occurrence_key] += 1
            source_seed = (
                f"{occurrence_key}\x1f{occurrences[occurrence_key]}"
            ).encode("utf-8")
            source_key = f"pnl:{hashlib.sha256(source_seed).hexdigest()[:32]}"
            parsed.append(
                ParsedTransaction(
                    source_key=source_key,
                    source_row=row_number,
                    kind=kind,
                    txn_date=txn_date,
                    voucher_number=None,
                    voucher_type="Profit & Loss summary",
                    party_name=None,
                    category=label,
                    gross_amount=amount,
                    net_amount=amount,
                    tax_amount=Decimal("0"),
                    lines=(
                        ParsedLine(
                            line_type="category",
                            name=label,
                            amount=amount,
                        ),
                    ),
                )
            )

    if not parsed:
        raise ImportValidationError(
            "The Profit & Loss sheet was recognized, but no supported ledger amounts "
            "could be mapped. Include ledger names beside numeric values, then try again."
        )
    if len({transaction.kind for transaction in parsed}) < 2:
        raise ImportValidationError(
            "The Profit & Loss sheet needs recognizable income and cost/expense sections. "
            "ARQ found only one side and stopped rather than showing a misleading result."
        )
    return parsed, skipped, len(set(period_by_column.values()))


def _profit_loss_kind(label: str) -> str | None:
    if _PROFIT_LOSS_PURCHASE.search(label):
        return "purchase"
    if _PROFIT_LOSS_SALES.search(label):
        return "sales"
    if _PROFIT_LOSS_EXPENSE.search(label):
        return "expense"
    return None


def _nearest_section(sections: dict[int, str], column: int) -> str | None:
    if column in sections:
        return sections[column]
    candidates = [
        (abs(section_column - column), kind)
        for section_column, kind in sections.items()
    ]
    return min(candidates)[1] if candidates else None


def _profit_loss_periods(sheet) -> dict[int, date]:
    periods: dict[int, date] = {}
    for values in sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row or 1, 15),
        max_col=min(sheet.max_column or 1, MAX_SHEET_COLUMNS),
        values_only=True,
    ):
        for column, value in enumerate(values):
            period = _period_date(value)
            if period:
                periods[column] = period
    return periods


def _period_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Numeric P&L amounts can also be valid Excel date serials. Only typed date
    # cells or explicit date text are treated as period headers.
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return None
    text = str(value or "").strip()
    if not text:
        return None
    direct = _date_value(text)
    if direct:
        return direct
    matches = re.findall(
        r"\d{1,2}[-/. ](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)"
        r"[a-z]*[-/. ]\d{2,4}",
        text,
        flags=re.I,
    )
    for candidate in reversed(matches):
        normalized = re.sub(r"[/. ]", "-", candidate)
        parsed = _date_value(normalized)
        if parsed:
            return parsed
    for pattern in ("%b %Y", "%B %Y", "%b-%Y", "%B-%Y", "%b %y", "%B %y"):
        try:
            parsed = datetime.strptime(text, pattern)
            return date(parsed.year, parsed.month, 1)
        except ValueError:
            continue
    return None


def _looks_like_amount(value: object) -> bool:
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return True
    text = str(value or "").strip()
    if not text:
        return False
    normalized = (
        text.replace(",", "")
        .replace("₹", "")
        .replace("Rs.", "")
        .replace("Rs", "")
        .strip()
    )
    return bool(re.fullmatch(r"\(?-?\d+(?:\.\d+)?\)?", normalized))


_FLAT_HEADER_ALIASES = {
    "date": ("DATE", "VOUCHERDATE", "INVOICEDATE", "BILLDATE"),
    "particulars": (
        "PARTICULARS",
        "ITEM",
        "ITEMNAME",
        "STOCKITEM",
        "STOCKITEMNAME",
        "PRODUCT",
        "PRODUCTNAME",
        "DESCRIPTION",
    ),
    "party": (
        "BUYER",
        "CUSTOMER",
        "CUSTOMERNAME",
        "PARTY",
        "PARTYNAME",
        "PARTYLEDGERNAME",
        "SUPPLIER",
        "SUPPLIERNAME",
        "VENDOR",
        "VENDORNAME",
    ),
    "voucher_type": ("VOUCHERTYPE", "VCHTYPE", "TYPE"),
    "voucher_number": (
        "VOUCHERNUMBER",
        "VOUCHERNO",
        "VCHNO",
        "INVOICENUMBER",
        "INVOICENO",
        "BILLNUMBER",
        "BILLNO",
        "REFERENCE",
        "REFNO",
    ),
    "quantity": (
        "QUANTITY",
        "QTY",
        "BILLEDQTY",
        "ACTUALQTY",
        "ALTUNITS",
        "ALTERNATEUNITS",
    ),
    "unit": ("UNIT", "UOM", "UNITOFMEASURE"),
    "rate": ("RATE", "PRICE", "UNITPRICE", "SELLINGRATE", "PURCHASERATE"),
    "net": (
        "VALUE",
        "NETVALUE",
        "NETAMOUNT",
        "TAXABLEVALUE",
        "TAXABLEAMOUNT",
        "AMOUNT",
        "LINEVALUE",
        "TOTAL",
    ),
    "gross": (
        "GROSSTOTAL",
        "GROSSVALUE",
        "GROSSAMOUNT",
        "INVOICEVALUE",
        "INVOICEAMOUNT",
        "TOTALAMOUNT",
    ),
}
_TOTAL_ROW = re.compile(r"^(grand\s+)?total\b", re.I)


def _find_flat_register(workbook):
    """Find a single-sheet register even when its detail rows change layout."""
    best = None
    for sheet in workbook.worksheets:
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 25), values_only=True),
            start=1,
        ):
            if len(values) > MAX_SHEET_COLUMNS:
                continue
            mapping = _flat_column_mapping(values)
            if mapping is None:
                continue
            score = 3 + sum(
                bool(mapping.get(role))
                for role in ("party", "voucher_type", "voucher_number", "rate", "gross")
            )
            candidate = (score, sheet, row_number, mapping)
            if best is None or candidate[0] > best[0]:
                best = candidate
    if best is None:
        return None
    _, sheet, header_row, mapping = best
    mapping["display"]["sheet"] = sheet.title
    return sheet, header_row, mapping


def _flat_column_mapping(headers: tuple[object, ...]) -> dict | None:
    normalized = [_header_key(value) for value in headers]

    def first(role: str) -> int | None:
        for alias in _FLAT_HEADER_ALIASES[role]:
            if alias in normalized:
                return normalized.index(alias)
        return None

    date_index = first("date")
    particulars_index = first("particulars")
    net_index = first("net")
    if date_index is None or particulars_index is None or net_index is None:
        return None

    quantity_indices = [
        index
        for index, key in enumerate(normalized)
        if key in _FLAT_HEADER_ALIASES["quantity"]
    ]
    tax_indices = [
        index
        for index, value in enumerate(headers)
        if _is_tax_ledger(_text(value))
    ]
    mapping = {
        "date": date_index,
        "particulars": particulars_index,
        "party": first("party"),
        "voucher_type": first("voucher_type"),
        "voucher_number": first("voucher_number"),
        "quantity": quantity_indices,
        "unit": first("unit"),
        "rate": first("rate"),
        "net": net_index,
        "gross": first("gross"),
        "tax": tax_indices,
    }

    def label(index: int | None) -> str | None:
        return _text(headers[index], 80) if index is not None else None

    display = {
        "sheet": "Detected table",
        "date": label(mapping["date"]),
        "product": label(mapping["particulars"]),
        "party": label(mapping["party"]) or label(mapping["particulars"]),
        "voucher_number": label(mapping["voucher_number"]) or "Not supplied",
        "quantity": " / ".join(
            label(index) or "" for index in quantity_indices
        ) or "Not supplied",
        "rate": label(mapping["rate"]) or "Not supplied",
        "net_value": label(mapping["net"]),
        "gross_value": label(mapping["gross"]) or "Not supplied",
        "tax": " / ".join(label(index) or "" for index in tax_indices) or "Not supplied",
    }
    mapping["display"] = display
    return mapping


def _classify_flat_register(
    sheet,
    header_row: int,
    mapping: dict,
    filename: str,
    declared_kind: str | None,
) -> tuple[str, float, str]:
    votes: Counter[str] = Counter()
    voucher_index = mapping.get("voucher_type")
    if voucher_index is not None:
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if row_number > MAX_SHEET_ROWS + header_row:
                raise ImportValidationError(
                    f"The '{sheet.title}' sheet exceeds the {MAX_SHEET_ROWS:,}-row limit."
                )
            if _date_value(_cell(values, mapping["date"])) is None:
                continue
            label = (_text(_cell(values, voucher_index)) or "").casefold()
            if re.search(r"\bsales?\b", label):
                votes["sales"] += 1
            elif re.search(r"\bpurchases?\b", label):
                votes["purchase"] += 1
            elif re.search(
                r"\b(expense|expenses|payment|journal|repair|maintenance|rent)\b",
                label,
            ):
                votes["expense"] += 1

    voted = [kind for kind, count in votes.items() if count]
    if "sales" in voted and "purchase" in voted:
        raise ImportValidationError(
            "This register mixes sales and purchase vouchers. Export and upload each type "
            "as a separate file."
        )
    if voted:
        detected = max(votes, key=votes.get)
        confidence = 0.96
        reason = (
            f"Single-sheet register detected on '{sheet.title}'; dated voucher rows identify "
            f"{detected} transactions and product detail rows reconcile to their values."
        )
    elif declared_kind:
        detected = declared_kind
        confidence = 0.78
        reason = (
            f"Single-sheet register detected on '{sheet.title}'; the selected upload type "
            "resolved its custom or missing voucher labels."
        )
    else:
        detected = _filename_kind(filename)
        if detected is None:
            raise ImportValidationError(
                "The register columns are usable, but its transaction type is ambiguous. "
                "Choose Sales, Purchase or Expense before uploading."
            )
        confidence = 0.65
        reason = (
            f"Single-sheet register detected on '{sheet.title}'; the filename supplied the "
            "transaction type because voucher labels were unavailable."
        )

    if declared_kind and declared_kind != detected:
        raise ImportValidationError(
            f"This workbook looks like {detected}, but it was uploaded as {declared_kind}. "
            f"Use the {detected.title()} upload option."
        )
    return detected, confidence, reason


def _build_flat_register_transactions(
    sheet,
    header_row: int,
    mapping: dict,
    kind: str,
) -> tuple[list[ParsedTransaction], int, int, tuple[str, ...]]:
    groups: list[tuple[int, tuple[object, ...], list[tuple[int, tuple[object, ...]]]]] = []
    footer_sections: list[tuple[int, int, int, tuple[object, ...]]] = []
    section_start = 0
    current: tuple[int, tuple[object, ...], list[tuple[int, tuple[object, ...]]]] | None = None

    for row_number, values in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if row_number > MAX_SHEET_ROWS + header_row:
            raise ImportValidationError(
                f"The '{sheet.title}' sheet exceeds the {MAX_SHEET_ROWS:,}-row limit."
            )
        if len(values) > MAX_SHEET_COLUMNS:
            raise ImportValidationError(
                f"The '{sheet.title}' sheet has too many columns to process safely."
            )

        txn_date = _date_value(_cell(values, mapping["date"]))
        particulars = _text(_cell(values, mapping["particulars"]))
        if txn_date is not None:
            if current is not None:
                groups.append(current)
            current = (row_number, values, [])
            continue
        if current is None or not particulars:
            continue
        if _TOTAL_ROW.search(particulars):
            groups.append(current)
            current = None
            footer_sections.append((section_start, len(groups), row_number, values))
            section_start = len(groups)
            continue
        if _header_key(particulars) in _FLAT_HEADER_ALIASES["particulars"]:
            continue
        current[2].append((row_number, values))

    if current is not None:
        groups.append(current)

    parsed: list[ParsedTransaction] = []
    skipped = 0
    fingerprints: Counter[str] = Counter()
    for source_row, parent, detail_rows in groups:
        txn_date = _date_value(_cell(parent, mapping["date"]))
        party = _text(_cell(parent, mapping.get("party")))
        if not party:
            party = _text(_cell(parent, mapping["particulars"]))

        item_lines = tuple(
            line
            for _, detail in detail_rows
            if (line := _flat_item_line(detail, mapping)) is not None
        )
        parent_net = abs(_amount(_cell(parent, mapping["net"])))
        item_net = sum((line.amount for line in item_lines), Decimal("0"))
        net = item_net or parent_net
        tax_lines = tuple(
            ParsedLine(
                line_type="tax",
                name=_text(sheet.cell(header_row, index + 1).value) or "Tax",
                amount=abs(_amount(_cell(parent, index))),
            )
            for index in mapping["tax"]
            if _amount(_cell(parent, index))
        )
        tax = sum((line.amount for line in tax_lines), Decimal("0"))
        gross = abs(_amount(_cell(parent, mapping.get("gross"))))
        if not gross:
            gross = max(parent_net, net + tax)
        if not net and gross:
            net = max(gross - tax, Decimal("0"))
        if gross <= 0:
            skipped += 1
            continue

        voucher_number = _text(_cell(parent, mapping.get("voucher_number")))
        raw_voucher_type = _text(_cell(parent, mapping.get("voucher_type")))
        voucher_type = (
            raw_voucher_type
            if raw_voucher_type and re.search(rf"\b{re.escape(kind.rstrip('s'))}s?\b", raw_voucher_type, re.I)
            else kind.title()
        )
        lines = item_lines + tax_lines
        category = max(item_lines, key=lambda line: line.amount).name if item_lines else None

        fingerprint = _flat_business_fingerprint(
            kind=kind,
            txn_date=txn_date,
            party=party,
            gross=gross,
            net=net,
            tax=tax,
            lines=item_lines,
        )
        fingerprints[fingerprint] += 1
        if voucher_number:
            source_key = _source_key(
                kind=kind,
                guid=None,
                sequence=str(source_row),
                txn_date=txn_date,
                voucher_number=voucher_number,
                party=party,
            )
        else:
            source_key = f"flat:{fingerprint}:{fingerprints[fingerprint]}"

        parsed.append(
            ParsedTransaction(
                source_key=source_key,
                source_row=source_row,
                kind=kind,
                txn_date=txn_date,
                voucher_number=voucher_number,
                voucher_type=voucher_type,
                party_name=party,
                category=category,
                gross_amount=_money(gross),
                net_amount=_money(net),
                tax_amount=_money(tax),
                lines=lines,
            )
        )

    # A real voucher number is authoritative. If a customized register repeats
    # it, keep only the last normalized copy so its lines cannot double-count.
    unique = {transaction.source_key: transaction for transaction in parsed}
    skipped += len(parsed) - len(unique)
    duplicate_groups = (
        sum(1 for count in fingerprints.values() if count > 1)
        if mapping.get("voucher_number") is None
        else 0
    )
    reconciliation_warnings = []
    for start, finish, footer_row, footer in footer_sections:
        stated_total = abs(_amount(_cell(footer, mapping["net"])))
        visible_total = sum(
            (
                abs(_amount(_cell(parent, mapping["net"])))
                for _, parent, _ in groups[start:finish]
            ),
            Decimal("0"),
        )
        tolerance = max(Decimal("1"), stated_total * Decimal("0.005"))
        if stated_total and abs(stated_total - visible_total) > tolerance:
            reconciliation_warnings.append(
                f"The total at row {footer_row} does not reconcile to visible voucher rows "
                f"(stated {_money(stated_total)}, visible {_money(visible_total)}). "
                "Only visible transactions were imported."
            )
    return (
        list(unique.values()),
        skipped,
        duplicate_groups,
        tuple(reconciliation_warnings),
    )


def _flat_item_line(values: tuple[object, ...], mapping: dict) -> ParsedLine | None:
    name = _text(_cell(values, mapping["particulars"]))
    if not name or _TOTAL_ROW.search(name):
        return None
    amount = abs(_amount(_cell(values, mapping["net"])))
    rate = _positive_decimal(_cell(values, mapping.get("rate")))
    quantity, quantity_unit = _best_flat_quantity(
        values,
        mapping["quantity"],
        rate,
        amount,
    )
    unit = _text(_cell(values, mapping.get("unit")), 30) or quantity_unit
    if not amount and quantity and rate:
        amount = quantity * rate
    if not amount:
        return None
    return ParsedLine(
        line_type="item",
        name=name,
        amount=_money(amount),
        quantity=quantity,
        unit=unit,
        rate=rate,
    )


def _best_flat_quantity(
    values: tuple[object, ...],
    indices: list[int],
    rate: Decimal | None,
    amount: Decimal,
) -> tuple[Decimal | None, str | None]:
    candidates = []
    for index in indices:
        quantity, unit = _quantity(_cell(values, index))
        if quantity is None:
            continue
        error = Decimal("1")
        if rate and amount:
            error = abs((quantity * rate) - amount) / max(amount, Decimal("0.01"))
        candidates.append((error, quantity, unit))
    if not candidates:
        return None, None
    _, quantity, unit = min(candidates, key=lambda item: item[0])
    return quantity, unit


def _flat_business_fingerprint(
    *,
    kind: str,
    txn_date: date | None,
    party: str | None,
    gross: Decimal,
    net: Decimal,
    tax: Decimal,
    lines: tuple[ParsedLine, ...],
) -> str:
    canonical_lines = sorted(
        [
            {
                "name": line.name.casefold(),
                "amount": str(_money(line.amount)),
                "quantity": str(line.quantity.normalize()) if line.quantity else "",
                "unit": (line.unit or "").casefold(),
                "rate": str(line.rate.normalize()) if line.rate else "",
            }
            for line in lines
        ],
        key=lambda line: json.dumps(line, sort_keys=True),
    )
    canonical = {
        "kind": kind,
        "date": txn_date.isoformat() if txn_date else "",
        "party": (party or "").casefold(),
        "gross": str(_money(gross)),
        "net": str(_money(net)),
        "tax": str(_money(tax)),
        "lines": canonical_lines,
    }
    payload = json.dumps(canonical, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _cell(values: tuple[object, ...], index: int | None) -> object:
    if index is None or index < 0 or index >= len(values):
        return None
    return values[index]


def _classify(
    filename: str,
    vouchers: list[dict[str, object]],
    ledgers: list[dict[str, object]],
    declared_kind: str | None,
) -> tuple[str, float, str]:
    voucher_types = Counter()
    for record in [*vouchers, *ledgers]:
        value = _text(
            _first(record, "VCHTYPE", "VOUCHERTYPENAME", "VOUCHERTYPE")
        )
        if value:
            voucher_types[value.casefold()] += 1

    joined_types = " ".join(voucher_types)
    sales_votes = sum(
        count for value, count in voucher_types.items() if re.search(r"\bsales?\b", value)
    )
    purchase_votes = sum(
        count for value, count in voucher_types.items() if re.search(r"\bpurchases?\b", value)
    )
    expense_votes = sum(
        count
        for value, count in voucher_types.items()
        if re.search(r"\b(expense|payment|journal|repair|maintenance|rent)\b", value)
    )

    if sales_votes and purchase_votes:
        raise ImportValidationError(
            "This workbook mixes sales and purchase vouchers. Export and upload each type "
            "as a separate file."
        )
    if sales_votes:
        detected = "sales"
        confidence = 0.99
        reason = "Tally voucher types identify sales transactions."
    elif purchase_votes:
        detected = "purchase"
        confidence = 0.99
        reason = "Tally voucher types identify purchase transactions."
    elif expense_votes:
        detected = "expense"
        confidence = 0.90
        reason = "Balanced journal/payment voucher entries identify expense transactions."
    elif declared_kind:
        detected = declared_kind
        confidence = 0.68
        reason = (
            "The workbook has a valid Tally ledger structure; the selected upload type "
            "resolved an otherwise custom voucher name."
        )
    else:
        filename_hint = _filename_kind(filename)
        if filename_hint:
            detected = filename_hint
            confidence = 0.60
            reason = (
                "The workbook has a valid Tally ledger structure and its filename supplied "
                "the only available transaction-type hint."
            )
        else:
            details = f" Found voucher types: {joined_types[:120]}." if joined_types else ""
            raise ImportValidationError(
                "The workbook structure is valid, but its transaction type is ambiguous. "
                "Choose Sales, Purchase or Expense before uploading." + details
            )

    if declared_kind and declared_kind != detected:
        raise ImportValidationError(
            f"This workbook looks like {detected}, but it was uploaded as {declared_kind}. "
            f"Use the {detected.title()} upload option."
        )
    return detected, confidence, reason


def _build_transactions(
    kind: str,
    vouchers: list[dict[str, object]],
    ledgers: list[dict[str, object]],
    inventory: list[dict[str, object]],
) -> tuple[list[ParsedTransaction], int]:
    voucher_meta = {
        _sequence(record.get("VOUCHERSEQ")): record
        for record in vouchers
        if _sequence(record.get("VOUCHERSEQ"))
    }
    ledger_groups = _group_by_sequence(ledgers)
    inventory_groups = _group_by_sequence(inventory)
    sequences = list(dict.fromkeys([*voucher_meta, *ledger_groups, *inventory_groups]))

    parsed: list[ParsedTransaction] = []
    skipped = 0
    for sequence in sequences:
        meta = voucher_meta.get(sequence, {})
        ledger_rows = ledger_groups.get(sequence, [])
        inventory_rows = inventory_groups.get(sequence, [])
        representative = ledger_rows[0] if ledger_rows else (
            inventory_rows[0] if inventory_rows else meta
        )

        party = _text(
            _first(meta, "PARTYLEDGERNAME", "PARTYNAME")
            or _first(representative, "PARTY", "PARTYNAME")
        )
        party_rows = [
            row
            for row in ledger_rows
            if _is_yes(row.get("ISPARTYLEDGER"))
            or (
                party
                and _text(row.get("LEDGERNAME"))
                and _text(row.get("LEDGERNAME")).casefold() == party.casefold()
            )
        ]
        tax_rows = [
            row
            for row in ledger_rows
            if row not in party_rows and _is_tax_ledger(_text(row.get("LEDGERNAME")))
        ]
        category_rows = [
            row
            for row in ledger_rows
            if row not in party_rows
            and row not in tax_rows
            and not _ROUNDING_LEDGER.search(_text(row.get("LEDGERNAME")) or "")
            and _amount(row.get("AMOUNT")) != 0
        ]

        gross = _dominant_side(_amount(row.get("AMOUNT")) for row in party_rows)
        tax = sum((abs(_amount(row.get("AMOUNT"))) for row in tax_rows), Decimal("0"))

        item_lines = _inventory_lines(inventory_rows)
        category_lines = _category_lines(category_rows)
        if item_lines:
            net = sum((line.amount for line in item_lines), Decimal("0"))
            business_lines = item_lines
        else:
            net = _kind_side(kind, (_amount(row.get("AMOUNT")) for row in category_rows))
            business_lines = category_lines

        if not gross:
            gross = max(net + tax, _dominant_side(_amount(row.get("AMOUNT")) for row in ledger_rows))
        if not net and gross:
            net = max(gross - tax, Decimal("0"))
        if gross <= 0:
            skipped += 1
            continue

        tax_lines = tuple(
            ParsedLine(
                line_type="tax",
                name=_text(row.get("LEDGERNAME")) or "Tax",
                amount=abs(_amount(row.get("AMOUNT"))),
            )
            for row in tax_rows
            if _amount(row.get("AMOUNT"))
        )
        lines = tuple(business_lines) + tax_lines
        category = max(business_lines, key=lambda line: line.amount).name if business_lines else None

        voucher_number = _text(
            _first(meta, "VOUCHERNUMBER")
            or _first(representative, "VOUCHERNUMBER")
        )
        voucher_type = _text(
            _first(meta, "VCHTYPE", "VOUCHERTYPENAME")
            or _first(representative, "VOUCHERTYPE")
        )
        txn_date = _date_value(
            _first(meta, "DATE") or _first(representative, "DATE")
        )
        guid = _text(_first(meta, "GUID") or _first(representative, "GUID"))
        source_key = _source_key(
            kind=kind,
            guid=guid,
            sequence=sequence,
            txn_date=txn_date,
            voucher_number=voucher_number,
            party=party,
        )

        parsed.append(
            ParsedTransaction(
                source_key=source_key,
                source_row=int(meta.get("_ROW") or representative.get("_ROW") or 0),
                kind=kind,
                txn_date=txn_date,
                voucher_number=voucher_number,
                voucher_type=voucher_type,
                party_name=party,
                category=category,
                gross_amount=_money(gross),
                net_amount=_money(net),
                tax_amount=_money(tax),
                lines=lines,
            )
        )
    # Some customized exports repeat the same voucher in more than one
    # sequence. Keep one normalized voucher so its item/category lines cannot
    # double the dashboard breakdown even though the transaction upsert is safe.
    unique = {transaction.source_key: transaction for transaction in parsed}
    skipped += len(parsed) - len(unique)
    return list(unique.values()), skipped


def _inventory_lines(rows: list[dict[str, object]]) -> tuple[ParsedLine, ...]:
    lines = []
    for row in rows:
        name = _text(row.get("STOCKITEMNAME"))
        amount = abs(_amount(row.get("AMOUNT")))
        if not name or not amount:
            continue
        quantity, unit = _quantity(row.get("BILLEDQTY") or row.get("ACTUALQTY"))
        lines.append(
            ParsedLine(
                line_type="item",
                name=name,
                amount=amount,
                quantity=quantity,
                unit=unit,
                rate=_positive_decimal(row.get("RATE")),
            )
        )
    return tuple(lines)


def _category_lines(rows: list[dict[str, object]]) -> tuple[ParsedLine, ...]:
    return tuple(
        ParsedLine(
            line_type="category",
            name=_text(row.get("LEDGERNAME")) or "Uncategorised",
            amount=abs(_amount(row.get("AMOUNT"))),
        )
        for row in rows
        if _amount(row.get("AMOUNT"))
    )


def _group_by_sequence(
    rows: list[dict[str, object]],
) -> dict[str, list[dict[str, object]]]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        sequence = _sequence(row.get("VOUCHERSEQ"))
        if sequence:
            grouped[sequence].append(row)
    return grouped


def _source_key(
    *,
    kind: str,
    guid: str | None,
    sequence: str,
    txn_date: date | None,
    voucher_number: str | None,
    party: str | None,
) -> str:
    if guid:
        return f"guid:{guid.strip().casefold()}"
    # VoucherSeq is an export-row locator and may change when the same period is
    # exported with a different sort order. Prefer business identity so a
    # modified re-export updates the existing voucher.
    voucher_identity = [
        kind,
        txn_date.isoformat() if txn_date else "",
        (voucher_number or "").strip().casefold(),
        (party or "").strip().casefold(),
    ]
    if any(voucher_identity[1:]):
        return "derived:" + "\x1f".join(voucher_identity)
    raw = "|".join(
        [kind, "sequence", sequence]
    )
    return "derived:" + hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _dominant_side(values: Iterable[Decimal]) -> Decimal:
    positive = Decimal("0")
    negative = Decimal("0")
    for value in values:
        if value > 0:
            positive += value
        elif value < 0:
            negative += abs(value)
    return max(positive, negative)


def _kind_side(kind: str, values: Iterable[Decimal]) -> Decimal:
    materialized = list(values)
    if kind == "sales":
        positive = sum((value for value in materialized if value > 0), Decimal("0"))
        if positive:
            return positive
    else:
        negative = sum((abs(value) for value in materialized if value < 0), Decimal("0"))
        if negative:
            return negative
    return _dominant_side(materialized)


def _header_key(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _normalize_kind(value: str | None) -> str | None:
    if value is None or not value.strip():
        return None
    normalized = value.strip().casefold()
    aliases = {
        "auto": None,
        "smart": None,
        "sale": "sales",
        "sales": "sales",
        "purchase": "purchase",
        "purchases": "purchase",
        "expense": "expense",
        "expenses": "expense",
        "profit_loss": "profit_loss",
        "profit loss": "profit_loss",
        "p&l": "profit_loss",
        "pnl": "profit_loss",
    }
    if normalized not in aliases:
        raise ImportValidationError(
            "Upload type must be auto, sales, purchase, expense or profit & loss."
        )
    kind = aliases[normalized]
    return kind


def _filename_kind(filename: str) -> str | None:
    stem = PurePath(filename).stem.casefold()
    if re.search(r"\bsales?\b", stem):
        return "sales"
    if re.search(r"\bpurchases?\b", stem):
        return "purchase"
    if re.search(r"\b(expense|expenses|expences)\b", stem):
        return "expense"
    return None


def _filename_is_profit_loss(filename: str) -> bool:
    stem = PurePath(filename).stem.casefold()
    return bool(
        re.search(r"\bprofit[\s_-]*(?:and|&)?[\s_-]*loss\b", stem)
        or re.search(r"\bp[\s_-]*&[\s_-]*l\b", stem)
        or re.search(r"\bpnl\b", stem)
        or re.search(r"\bincome[\s_-]*statement\b", stem)
    )


def _sequence(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _first(record: dict[str, object], *keys: str) -> object:
    for key in keys:
        value = record.get(key)
        if value not in (None, ""):
            return value
    return None


def _text(value: object, max_length: int = 300) -> str | None:
    if value in (None, ""):
        return None
    cleaned = " ".join(str(value).split()).strip()
    return cleaned[:max_length] if cleaned else None


def _is_yes(value: object) -> bool:
    return str(value or "").strip().casefold() in {"yes", "true", "1", "y"}


def _is_tax_ledger(name: str | None) -> bool:
    return bool(name and _TAX_LEDGER.search(name))


def _amount(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "").replace("₹", "").replace("Rs.", "")
    match = _NUMBER.search(text)
    if not match:
        return Decimal("0")
    try:
        amount = Decimal(match.group())
    except InvalidOperation:
        return Decimal("0")
    if text.startswith("(") and text.endswith(")"):
        amount = -abs(amount)
    return amount


def _positive_decimal(value: object) -> Decimal | None:
    parsed = abs(_amount(value))
    return parsed or None


def _quantity(value: object) -> tuple[Decimal | None, str | None]:
    if value in (None, ""):
        return None, None
    text = str(value).strip().replace(",", "")
    match = _NUMBER.search(text)
    if not match:
        return None, _text(text, 30)
    try:
        quantity = abs(Decimal(match.group()))
    except InvalidOperation:
        return None, _text(text, 30)
    unit = _text(text[match.end() :].strip(), 30)
    return quantity, unit


def _date_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 1 <= value <= 2_958_465:
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except (OverflowError, ValueError):
            return None
    text = str(value or "").strip()
    for pattern in (
        "%Y%m%d",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%d-%b-%Y",
        "%d-%b-%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d.%m.%Y",
        "%d.%m.%y",
    ):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"))
